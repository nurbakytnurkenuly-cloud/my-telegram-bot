import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import database


def generate_excel_report():
    # 1. Дерекқорды жүктейміз
    db = database.load_db()

    # 2. Мәліметтерді тізімге жинаймыз
    data_list = []
    for uid, info in db.items():
        data_list.append({
            "Telegram ID": uid,
            "Аты-жөні": info.get("name", "Белгісіз"),
            "Қаласы": info.get("city", "Белгісіз"),
            "Курсы": info.get("course", "Белгісіз"),
            "Тіркелген күні": info.get("reg_date", "-"),
            "Барлық сұрақ": info.get("total", 0),
            "Дұрыс жауап": info.get("correct", 0)
        })

    # 3. DataFrame құру
    df = pd.DataFrame(data_list)
    file_path = "Оқушылар_статистикасы.xlsx"

    # 4. Excel-ге бастапқы деректерді сақтау
    df.to_excel(file_path, index=False)

    # 5. Стильдеу үшін файлды қайта ашу
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Стиль параметрлері
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Шапканы бояу және ортаға келтіру
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Бағандар енін автоматты түрде реттеу
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Баған атауы (A, B, C...)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 6. Өзгерістерді сақтау (БҰЛ ӨТЕ МАҢЫЗДЫ!)
    wb.save(file_path)
    print(f"✅ Файл сәтті жасалды және безендірілді: {file_path}")


if __name__ == "__main__":
    generate_excel_report()